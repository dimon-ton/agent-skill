#!/usr/bin/env python3
"""
Extract schedule data from Phontan school Excel files.
Handles multiple sheets, grades, and output formats.
Works with any actual sheet in the Excel file.
"""

import openpyxl
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional, Any

class PhontanScheduleReader:
    """Read and parse Phontan school schedule Excel files."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        self.wb = openpyxl.load_workbook(file_path)
        self.all_sheets = self.wb.sheetnames

    def list_sheets(self) -> List[str]:
        """List all available sheets in the workbook."""
        return self.all_sheets

    def get_sheet(self, sheet_name: Optional[str] = None) -> openpyxl.worksheet.worksheet.Worksheet:
        """Get worksheet by name, with fuzzy matching support."""
        if sheet_name is None:
            # Default: use the first schedule sheet (not metadata sheets)
            for s in self.all_sheets:
                # Skip non-schedule sheets
                if s in ['วิชา', 'Sheet1', 'personWithSubject', 'colorTeacher', 'รายคน']:
                    continue
                if 'เทอม' in s or 'ปี' in s or s.startswith(('ป', 'ม')):
                    return self.wb[s]
            # Fallback to first sheet
            return self.wb[self.all_sheets[0]]

        # Exact match first
        if sheet_name in self.all_sheets:
            return self.wb[sheet_name]

        # Fuzzy match (case-insensitive substring)
        for s in self.all_sheets:
            if sheet_name.lower() in s.lower():
                return self.wb[s]

        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets:\n" +
                        "\n".join(f"  - {s}" for s in self.all_sheets))

    def extract_schedule(self, sheet_name: Optional[str] = None,
                        teacher_name: Optional[str] = None,
                        grade: Optional[str] = None,
                        day: Optional[str] = None,
                        row_range: Optional[tuple] = None) -> Dict:
        """
        Extract schedule data with optional filtering.

        Args:
            sheet_name: Name of the sheet (auto-detected if None)
            teacher_name: Filter by teacher name (substring match)
            grade: Filter by grade (e.g., 'ป.4', 'ม.1')
            day: Filter by day (Thai name)
            row_range: Tuple of (start_row, end_row) to extract

        Returns:
            Dictionary with structured schedule data
        """
        sheet = self.get_sheet(sheet_name)

        if row_range is None:
            row_range = (1, 100)

        # Extract all cell values from the range
        all_values = []
        for row in sheet.iter_rows(min_row=row_range[0], max_row=row_range[1], values_only=True):
            all_values.append(list(row))

        # Parse schedule structure
        schedule_data = {
            "sheet_name": sheet.title,
            "raw_data": all_values,  # Keep raw data for debugging
            "by_teacher": {},
            "by_day": {},
            "by_grade": {}
        }

        # Days of week in Thai
        thai_days = ['จันทร์', 'อังคาร', 'พุธ', 'พฤหัสบดี', 'ศุกร์']

        # Parse the table: look for patterns
        current_day = None
        current_grade = None
        current_row_idx = 0

        for row_idx, row in enumerate(all_values):
            for col_idx, cell_val in enumerate(row):
                if not cell_val:
                    continue

                cell_str = str(cell_val).strip()

                # Detect day
                if cell_str in thai_days:
                    current_day = cell_str
                    if current_day not in schedule_data["by_day"]:
                        schedule_data["by_day"][current_day] = []

                # Detect grade (ป.1, ป.2, ม.1, etc.)
                elif cell_str.startswith(('ป.', 'ม.')) and len(cell_str) <= 4:
                    current_grade = cell_str
                    if current_grade not in schedule_data["by_grade"]:
                        schedule_data["by_grade"][current_grade] = []

                # Detect teacher (starts with "ครู")
                elif cell_str.startswith('ครู'):
                    teacher = cell_str
                    if teacher not in schedule_data["by_teacher"]:
                        schedule_data["by_teacher"][teacher] = []

            current_row_idx = row_idx

        # Apply filters
        filtered = self._filter_results(schedule_data, teacher_name, grade, day)
        return filtered

    def _filter_results(self, data: Dict, teacher: Optional[str] = None,
                       grade: Optional[str] = None, day: Optional[str] = None) -> Dict:
        """Filter results by teacher, grade, or day."""
        filtered = {
            "sheet_name": data["sheet_name"],
            "by_teacher": {},
            "by_day": {},
            "by_grade": {}
        }

        # Filter by teacher
        if teacher:
            for t in data["by_teacher"]:
                if teacher.lower() in t.lower():
                    filtered["by_teacher"][t] = data["by_teacher"][t]
        else:
            filtered["by_teacher"] = data["by_teacher"]

        # Filter by day
        if day:
            if day in data["by_day"]:
                filtered["by_day"][day] = data["by_day"][day]
        else:
            filtered["by_day"] = data["by_day"]

        # Filter by grade
        if grade:
            if grade in data["by_grade"]:
                filtered["by_grade"][grade] = data["by_grade"][grade]
        else:
            filtered["by_grade"] = data["by_grade"]

        return filtered

    def search_teacher(self, sheet_name: Optional[str] = None,
                      teacher_name: Optional[str] = None) -> List[Dict]:
        """Search for teacher and return all their classes."""
        sheet = self.get_sheet(sheet_name)

        results = []
        thai_days = ['จันทร์', 'อังคาร', 'พุธ', 'พฤหัสบดี', 'ศุกร์']

        current_day = None
        current_grade = None
        current_time = None
        current_subject = None

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100, values_only=True)):
            for col_idx, cell_val in enumerate(row):
                if not cell_val:
                    continue

                cell_str = str(cell_val).strip()

                # Track day
                if cell_str in thai_days:
                    current_day = cell_str

                # Track grade
                elif cell_str.startswith(('ป.', 'ม.')) and len(cell_str) <= 4:
                    current_grade = cell_str

                # Track time (format: HH.MM-HH.MM)
                elif '.' in cell_str and '-' in cell_str and len(cell_str) == 13:
                    current_time = cell_str

                # Track subject and find teacher
                elif cell_str.startswith('ครู'):
                    if teacher_name and teacher_name.lower() in cell_str.lower():
                        results.append({
                            "day": current_day,
                            "grade": current_grade,
                            "time": current_time,
                            "subject": current_subject,
                            "teacher": cell_str,
                            "row": row_idx + 1
                        })
                # Subject (anything else that's not recognized pattern)
                elif len(cell_str) > 2 and not cell_str.startswith(('ป.', 'ม.', 'ครู', 'จ', 'อ', 'พ', 'ศ')):
                    current_subject = cell_str

        return results

    def to_table(self, data: Dict) -> str:
        """Convert schedule data to formatted table."""
        lines = [f"\nSheet: {data['sheet_name']}\n"]
        lines.append("=" * 100)

        if data.get("by_teacher"):
            lines.append("\n--- BY TEACHER ---")
            lines.append(f"{'Teacher':<25} {'Count':<8}")
            lines.append("-" * 100)
            for teacher, items in data["by_teacher"].items():
                lines.append(f"{teacher:<25} {len(items) if items else 0:<8}")

        if data.get("by_day"):
            lines.append("\n--- BY DAY ---")
            for day, items in data["by_day"].items():
                lines.append(f"{day:<25} {len(items) if items else 0:<8} classes")

        if data.get("by_grade"):
            lines.append("\n--- BY GRADE ---")
            for grade, items in data["by_grade"].items():
                lines.append(f"{grade:<25} {len(items) if items else 0:<8} classes")

        lines.append("=" * 100)
        return "\n".join(lines)

    def to_json(self, data: Dict, indent: int = 2) -> str:
        """Convert schedule data to JSON."""
        # Remove raw_data for cleaner JSON output
        output = {k: v for k, v in data.items() if k != "raw_data"}
        return json.dumps(output, ensure_ascii=False, indent=indent)


def main():
    """CLI interface for the schedule reader."""
    if len(sys.argv) < 2:
        print("Usage: python extract_schedule.py <file_path> [options]")
        print("Options:")
        print("  --list-sheets          List all available sheets")
        print("  --sheet <name>         Specify sheet name")
        print("  --teacher <name>       Search for teacher")
        print("  --grade <grade>        Filter by grade (e.g., ป.4, ม.1)")
        print("  --day <day>            Filter by day (Thai name)")
        print("  --format json|table    Output format (default: table)")
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        reader = PhontanScheduleReader(file_path)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Parse arguments
    list_sheets_flag = False
    sheet_name = None
    teacher_name = None
    grade = None
    day = None
    output_format = "table"

    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == "--list-sheets":
            list_sheets_flag = True
            i += 1
        elif sys.argv[i] == "--sheet" and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--teacher" and i + 1 < len(sys.argv):
            teacher_name = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--grade" and i + 1 < len(sys.argv):
            grade = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--day" and i + 1 < len(sys.argv):
            day = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--format" and i + 1 < len(sys.argv):
            output_format = sys.argv[i + 1]
            i += 2
        else:
            i += 1

    # Handle list sheets
    if list_sheets_flag:
        print("Available sheets:")
        for idx, sheet in enumerate(reader.list_sheets(), 1):
            print(f"  {idx}. {sheet}")
        return

    # Search for teacher
    if teacher_name:
        results = reader.search_teacher(sheet_name, teacher_name)
        if results:
            print(f"Found {len(results)} classes for {teacher_name}:")
            print("=" * 100)
            print(f"{'Day':<15} {'Grade':<8} {'Time':<15} {'Subject':<25} {'Teacher':<20}")
            print("-" * 100)
            for r in results:
                day = r.get('day') or 'N/A'
                grade = r.get('grade') or 'N/A'
                time_slot = r.get('time') or 'N/A'
                subject = r.get('subject') or 'N/A'
                teacher = r.get('teacher') or 'N/A'
                print(f"{day:<15} {grade:<8} {time_slot:<15} {subject:<25} {teacher:<20}")
            print("=" * 100)
        else:
            print(f"No classes found for {teacher_name}")
        return

    # Extract schedule
    data = reader.extract_schedule(sheet_name, teacher_name, grade, day)

    if output_format == "json":
        print(reader.to_json(data))
    else:
        print(reader.to_table(data))


if __name__ == "__main__":
    main()
